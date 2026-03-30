# import openpyxl
# import os
#
# # Replace "folder_path" with the path to your folder containing Excel files
# folder_path = r"path/to/folder" #insert the folder path between the ""
#
# # Loop through all Excel files in the folder
# for file_name in os.listdir(folder_path):
#     if file_name.endswith(".xlsx"): # Check if the file is an Excel file
#         file_path = os.path.join(folder_path, file_name)
#         print("Processing file:", file_path)
#
#         # Load the Excel file
#         workbook = openpyxl.load_workbook(file_path)
#
#         # Modify the worksheet(s)
#         for sheet in workbook.worksheets:
#             for row in sheet.iter_rows():
#                 for cell in row:
#                     if cell.value == "old_word": #insert the word it should find
#                         cell.value = "new_word" #insert the word it should replace it with
#
#         # Save the modified Excel file
#         workbook.save(file_path)
#
#         # Close the Excel file
#         workbook.close()
#
# import os
# import openpyxl
#
# folder_path = " #Insert Folder Path Here "
# excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')][:10]  # Test on 10 files
#
# replacement_text = """#Input Replacement Text Here"""
#
# for file_name in excel_files:
#     file_path = os.path.join(folder_path, file_name)
#     wb = openpyxl.load_workbook(file_path)
#
#     for sheet in wb:
#         for row in sheet.iter_rows(min_row=3, max_row=6):
#             for cell in row:
#                 cell.value = replacement_text
#
#     wb.save(file_path)
#     wb.close()
#
# print("All files processed and modified.")

#
# import os
# import openpyxl
#
# folder_path = "#Input Folder Path Here"
# excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')][:10]  # Test on 10 files
#
# replacement_texts = {
#Input Replacement Texts Here
# }
#
# for file_name in excel_files:
#     file_path = os.path.join(folder_path, file_name)
#     wb = openpyxl.load_workbook(file_path)
#
#     for sheet in wb:
#         for row in sheet.iter_rows(min_row=3, max_row=6):
#             for row_number, replacement_text in replacement_texts.items():
#                 row[row_number - 1][1].value = replacement_text
#                 for cell in row:
#                     if cell.column != 2:  # Clear other cells in the row
#                         cell.value = None
#
#     wb.save(file_path)
#     wb.close()
#
# print("All files processed and modified.")

import os
import openpyxl

# folder_path = "#Input Folder Path Here"
# excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')][:8]  # Test on 10 files
#
# replacement_texts = {
#Input Replacement Texts Here
# }
#
# for file_name in excel_files:
#     file_path = os.path.join(folder_path, file_name)
#     wb = openpyxl.load_workbook(file_path)
#
#     for sheet in wb:
#         for row in sheet.iter_rows(min_row=3, max_row=6):
#             for row_number, replacement_text in replacement_texts.items():
#                 row[0].offset(row_number - 1, 1).value = replacement_text
#                 for cell in row:
#                     if cell != row[0].offset(row_number - 1, 1):  # Clear other cells in the row
#                         cell.value = None
#
#     wb.save(file_path)
#     wb.close()
#
# print("All files processed and modified.")

